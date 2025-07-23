from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy
import os
from datetime import datetime
import calendar
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.formatting.rule import FormulaRule

light_border = Border(
    left=Side(border_style="hair", color="CCCCCC"),
    right=Side(border_style="hair", color="CCCCCC"),
    top=Side(border_style="hair", color="CCCCCC"),
    bottom=Side(border_style="hair", color="CCCCCC")
)

dark_border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

def detect_and_replace_date_rows(ws):
    fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for row in range(1, 4):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    found_date = None
    for row in [3, 2]:
        cell = ws.cell(row=row, column=1)
        val = cell.value
        if isinstance(val, datetime):
            found_date = val
            break
        elif isinstance(val, str):
            try:
                found_date = datetime.strptime(val.strip(), "%A %d %B %Y")
                break
            except:
                try:
                    found_date = datetime.strptime(val.strip(), "%d/%m/%Y")
                    break
                except:
                    pass
    if found_date:
        for rng in list(ws.merged_cells.ranges):
            if (rng.min_row <= 2 <= rng.max_row) or (rng.min_row <= 3 <= rng.max_row):
                ws.unmerge_cells(str(rng))
        for col in range(1, 12):
            ws.cell(row=2, column=col).value = None
            ws.cell(row=3, column=col).value = None
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=11)
        ws.cell(row=3, column=1).value = found_date
        ws.cell(row=3, column=1).number_format = "DD/MM/YYYY"
    for row in range(1, 4):
        for col in range(1, 12):
            ws.cell(row=row, column=col).fill = fill

def unmerge_vertical_and_fill(ws):
    merged_cells = list(ws.merged_cells.ranges)
    for merged_range in merged_cells:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_col == max_col and max_row > min_row:
            if 4 <= min_col <= 8:
                continue
            top_cell = ws.cell(row=min_row, column=min_col)
            top_value = top_cell.value
            ws.unmerge_cells(str(merged_range))
            for row in range(min_row, max_row + 1):
                target_cell = ws.cell(row=row, column=min_col)
                target_cell.value = top_value
                target_cell.font = copy(top_cell.font)

def find_row_by_text(ws, search_text):
    for row in range(1, ws.max_row + 1):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and search_text.lower() in cell.value.lower():
                return row
    return None

def clear_outside_columns(ws):
    for row in range(1, ws.max_row + 1):
        for col in range(12, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.border = Border()
            cell.font = Font()
            cell.fill = PatternFill()
            cell.alignment = Alignment()

def apply_light_borders(ws):
    for row in range(1, ws.max_row + 1):
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = light_border

def apply_dark_borders(ws, start_row, end_row, si_col=True):
    si_number = 1
    for row in range(start_row, end_row + 1):
        is_data_row = any(ws.cell(row=row, column=col).value not in [None, ""] for col in range(2, 12))
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            font = copy(cell.font)
            font.name = 'Calibri'
            font.size = 11
            cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = dark_border
        if si_col and is_data_row:
            ws.cell(row=row, column=1, value=si_number)
            si_number += 1

def set_row_heights(ws):
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 24

def autofit_columns(ws):
    for col in range(1, 12):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_length = max(max_length, len(str(val)))
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(col)].width = adjusted_width

def center_align_all(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def delete_irrelevant_hidden_sheets(wb, month_year_str):
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if wb.sheetnames.index(sheet) == 0:
            continue
        if ws.sheet_state == 'hidden' or ws.sheet_state == 'veryHidden':
            if month_year_str.lower() not in sheet.lower():
                wb.remove(ws)

def add_sales_summary_sheet(wb, month, year):
    for sheet_name in wb.sheetnames:
        if "sales summary" in sheet_name.lower():
            std = wb[sheet_name]
            wb.remove(std)

    ws = wb.create_sheet("Sales Summary " + month.upper(), 0)
    ws.sheet_view.zoomScale = 100

    default_font = Font(name='Calibri', size=12)
    bold_font = Font(name='Calibri', size=12, bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws['A1'] = "TARGET"
    ws['A1'].font = bold_font
    ws['A1'].fill = light_fill
    ws['B1'] = "AED2,000,000.00"
    ws['B1'].font = bold_font
    ws['B1'].fill = light_fill

    headers = ["Date", "Daily (AED)", "Banquet (AED)", "Total Sales (AED)", "CHECK"]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    month_num = datetime.strptime(month, "%b").month
    days_in_month = calendar.monthrange(year, month_num)[1]
    base_date = datetime(year, month_num, 1)

    for i in range(days_in_month):
        row = 4 + i
        date = base_date.replace(day=i + 1)
        ws.cell(row=row, column=1, value=date.strftime("%d/%m/%Y")).font = default_font

        daily_formula = f'''=IFERROR(INDEX(INDIRECT("'" & TEXT(A{row}, "mmm d") & "'!G:G"), MATCH(TRUE, INDEX(ISNUMBER(SEARCH("daily", INDIRECT("'" & TEXT(A{row}, "mmm d") & "'!D:D"))), 0), 0)), "")'''
        ws.cell(row=row, column=2, value=daily_formula).font = default_font

        banquet_formula = f'''=IFERROR(INDEX(INDIRECT("'" & TEXT(A{row}, "mmm d") & "'!G:G"), MATCH(TRUE, INDEX(ISNUMBER(SEARCH("banquet", INDIRECT("'" & TEXT(A{row}, "mmm d") & "'!D:D"))), 0), 0)), "")'''
        ws.cell(row=row, column=3, value=banquet_formula).font = default_font

        total_formula = f'''=IFERROR(INDEX(INDIRECT("'" & TEXT(A{row}, "mmm d") & "'!G:G"), MATCH("Grand Total:", INDIRECT("'" & TEXT(A{row}, "mmm d") & "'!E:E"), 0)), "")'''
        ws.cell(row=row, column=4, value=total_formula).font = default_font

        check_formula = f'''=IF(D{row}=B{row}+C{row}, "✅", "❌")'''
        ws.cell(row=row, column=5, value=check_formula).font = default_font

    total_row = 4 + days_in_month
    ws.cell(row=total_row + 2, column=1, value="TOTAL").font = bold_font
    ws.cell(row=total_row + 2, column=1).fill = light_fill
    ws.cell(row=total_row + 2, column=4, value=f"=SUM(D4:D{total_row - 1})").font = bold_font
    ws.cell(row=total_row + 2, column=4).fill = light_fill

    remaining_row = total_row + 3
    ws.cell(row=remaining_row, column=1, value="REMAINING").font = bold_font
    ws.cell(row=remaining_row, column=1).fill = light_fill
    ws.cell(row=remaining_row, column=4, value=f"=B1-D{total_row + 2}").font = bold_font
    ws.cell(row=remaining_row, column=4).fill = light_fill

    percent_row = remaining_row + 2
    ws.cell(row=percent_row, column=1, value="% ACHIEVED").font = bold_font
    ws.cell(row=percent_row, column=1).fill = light_fill
    ws.cell(row=percent_row, column=4, value=f"=D{total_row + 2}/B1").font = bold_font
    ws.cell(row=percent_row, column=4).fill = light_fill
    ws.cell(row=percent_row, column=4).number_format = FORMAT_PERCENTAGE_00

    bar_formula = f'=REPT("█", ROUND(D{percent_row}*20, 0)) & REPT("░", 20 - ROUND(D{percent_row}*20, 0)) & " " & TEXT(D{percent_row}, "0%")'
    bar_cell = ws.cell(row=percent_row, column=1, value=bar_formula)
    bar_cell.font = bold_font
    bar_cell.fill = light_fill
    bar_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.conditional_formatting.add(
        f"A{percent_row}",
        FormulaRule(formula=[f"$D${percent_row}<0.33"],
                    fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                    font=Font(color="9C0006"))
    )

    ws.conditional_formatting.add(
        f"A{percent_row}",
        FormulaRule(formula=[f"AND($D${percent_row}>=0.33,$D${percent_row}<0.66)"],
                    fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                    font=Font(color="9C6500"))
    )

    ws.conditional_formatting.add(
        f"A{percent_row}",
        FormulaRule(formula=[f"$D${percent_row}>=0.66"],
                    fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                    font=Font(color="006100"))
    )

    for row in ws.iter_rows(min_row=1, max_row=percent_row, min_col=1, max_col=5):
        for cell in row:
            if not cell.font:
                cell.font = default_font
            if cell.row < 4 or cell.row > (3 + days_in_month):
                cell.fill = light_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.column_dimensions['A'].width = 30
    for col_letter in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col_letter].width = 16

def rename_day_sheets(wb, month, year):
    month_num = datetime.strptime(month, "%b").month
    new_names = {}
    for sheet_name in wb.sheetnames:
        if "sales summary" in sheet_name.lower():
            continue
        ws = wb[sheet_name]
        try:
            day = int(''.join(filter(str.isdigit, sheet_name)))
            new_name = f"{month} {day}"
            if new_name in new_names.values():
                count = 1
                base_new_name = new_name
                while new_name in new_names.values():
                    count += 1
                    new_name = f"{base_new_name} ({count})"
            new_names[sheet_name] = new_name
        except:
            pass
    for old_name, new_name in new_names.items():
        ws = wb[old_name]
        ws.title = new_name

def process_workbook(file_stream):
    wb = openpyxl.load_workbook(file_stream)
    base_name = os.path.splitext(file_stream.name)[0]
    month_year_str = base_name.split()[-2].lower()
    year = int(base_name.split()[-1])
    month = base_name.split()[-2][:3].capitalize()

    delete_irrelevant_hidden_sheets(wb, month_year_str)
    add_sales_summary_sheet(wb, month, year)
    rename_day_sheets(wb, month, year)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if "sales summary" in sheet_name.lower():
            continue
        detect_and_replace_date_rows(ws)
        unmerge_vertical_and_fill(ws)
        clear_outside_columns(ws)
        apply_light_borders(ws)
        set_row_heights(ws)
        daily_row = find_row_by_text(ws, "Daily Catering Revenue")
        total_row = find_row_by_text(ws, "Grand Total:")
        if not daily_row or not total_row:
            if total_row:
                apply_dark_borders(ws, 1, total_row)
        else:
            apply_dark_borders(ws, 1, daily_row)
            apply_dark_borders(ws, daily_row + 2, total_row)
        center_align_all(ws)
        autofit_columns(ws)

    for sheet in wb.worksheets:
        sheet.sheet_state = 'visible'
        sheet.sheet_view.selection = [openpyxl.worksheet.views.Selection(activeCell="A1", sqref="A1")]
        sheet.sheet_view.sheetViewPane = None
        sheet.sheet_view.zoomScale = 100

    wb.active = 0
    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream